"""
export_excel.py
Generates a formatted Excel workbook from filtered KCET data.
One sheet per college + a summary sheet.
"""

import logging
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Colours
# ──────────────────────────────────────────────────────────────────────────────

NAVY       = "1A3A5C"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
ALT_ROW    = "F0F4F8"
HEADER_FG  = "FFFFFF"
BORDER_CLR = "C0CCD8"


def _thin_border():
    side = Side(style="thin", color=BORDER_CLR)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor=NAVY)


def _alt_fill():
    return PatternFill("solid", fgColor=ALT_ROW)


def _college_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE)


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(filtered_colleges: list[dict], title: str = "KCET Cutoff Analysis") -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ──
    _build_summary_sheet(wb, filtered_colleges, title)

    # ── Per-college sheets ──
    for college in filtered_colleges:
        df = college.get("table")
        if df is None or df.empty:
            continue
        sheet_name = _safe_sheet_name(college.get("college_code", "COL"))
        _build_college_sheet(wb, sheet_name, college, df)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ──────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, colleges: list[dict], title: str):
    ws = wb.create_sheet(title="Summary")

    # Title row
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
    title_cell.fill = _header_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = ["College Code", "College Name", "Matched Branches", "Branch Names"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 20

    # Data rows
    for i, college in enumerate(colleges, 3):
        df = college.get("table")
        course_col = df.columns[0] if df is not None and not df.empty else "Course Name"
        branches = list(df[course_col]) if df is not None and not df.empty else []
        branch_str = "\n".join(branches)
        n_branches = len(branches)

        row_data = [
            college.get("college_code", ""),
            college.get("college_name", ""),
            n_branches,
            branch_str,
        ]
        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col == 4),
            )

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 60


def _build_college_sheet(wb: Workbook, sheet_name: str, college: dict, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name)

    # College header row
    n_cols = len(df.columns)
    header_str = college.get("college_header", college.get("college_name", ""))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 1))
    h_cell = ws.cell(row=1, column=1, value=header_str)
    h_cell.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    h_cell.fill = _college_fill()
    h_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    h_cell.border = _thin_border()
    ws.row_dimensions[1].height = 28

    # Column headers (row 2)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=8)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 18

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        fill = _alt_fill() if row_idx % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col_idx, col_name in enumerate(df.columns, 1):
            val = str(row[col_name]) if str(row[col_name]) not in ("nan", "None") else "--"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=8)
            cell.fill = fill
            cell.border = _thin_border()
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    if df.columns[0]:
        ws.column_dimensions[get_column_letter(1)].width = 32
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    # Freeze panes below header
    ws.freeze_panes = "B3"


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """Excel sheet names max 31 chars, no special chars."""
    import re
    name = re.sub(r"[\\\/\*\?\:\[\]]", "", str(name))
    return name[:31]


def export_to_csv(filtered_colleges: list[dict]) -> BytesIO:
    """Export all filtered data as a single CSV."""
    frames = []
    for college in filtered_colleges:
        df = college.get("table")
        if df is None or df.empty:
            continue
        meta = pd.DataFrame(
            [[""] * len(df.columns)],
            columns=df.columns,
        )
        header_df = pd.DataFrame(
            [[college.get("college_header", "")] + [""] * (len(df.columns) - 1)],
            columns=df.columns,
        )
        frames.extend([header_df, df, meta])

    if not frames:
        buf = BytesIO()
        buf.write(b"No data found.\n")
        buf.seek(0)
        return buf

    combined = pd.concat(frames, ignore_index=True)
    buf = BytesIO()
    combined.to_csv(buf, index=False)
    buf.seek(0)
    return buf
